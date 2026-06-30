"""
report.py — Report generation: plots, Excel export, and console summary
"""
import logging
import os
import json
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─── Color palette ───────────────────────────────────────────
TEAL   = "1D9E75"
AMBER  = "BA7517"
CORAL  = "D85A30"
PURPLE = "534AB7"
GRAY   = "5F5E5A"
GREEN  = "3B6D11"
LIGHT_BG = "F1EFE8"
WHITE  = "FFFFFF"


def _hex(h: str) -> PatternFill:
    return PatternFill("solid", fgColor=h)


def _border() -> Border:
    s = Side(style="thin", color="D3D1C7")
    return Border(left=s, right=s, top=s, bottom=s)


# ─── Matplotlib plots ────────────────────────────────────────

def plot_actual_vs_predicted(y_true, y_pred, output_path: str, title: str = ""):
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    fig.suptitle(f"Actual vs Predicted ROP{' — ' + title if title else ''}", fontsize=13, fontweight="bold")
    
    # Scatter
    ax = axes[0]
    ax.scatter(y_true, y_pred, alpha=0.3, s=8, color="#1D9E75")
    lo = min(y_true.min(), y_pred.min())
    hi = max(y_true.max(), y_pred.max())
    ax.plot([lo, hi], [lo, hi], "r--", lw=1.5, label="Perfect prediction")
    ax.set_xlabel("Actual ROP (ft/hr)")
    ax.set_ylabel("Predicted ROP (ft/hr)")
    ax.set_title("Scatter: Actual vs Predicted")
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)
    
    # Residuals
    ax = axes[1]
    residuals = y_pred - y_true
    ax.hist(residuals, bins=60, color="#534AB7", alpha=0.75, edgecolor="white", linewidth=0.3)
    ax.axvline(0, color="red", linestyle="--", lw=1.5)
    ax.set_xlabel("Residual (ft/hr)")
    ax.set_ylabel("Count")
    ax.set_title("Residual Distribution")
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close()
    logger.info(f"Plot saved: {output_path}")


def plot_feature_importance(fi_df: pd.DataFrame, output_path: str):
    fig, ax = plt.subplots(figsize=(9, 5))
    colors = ["#1D9E75" if i < 3 else "#9FE1CB" for i in range(len(fi_df))]
    bars = ax.barh(fi_df["feature"][::-1], fi_df["importance"][::-1], color=colors[::-1])
    ax.set_xlabel("Importance (split count)")
    ax.set_title("Feature Importance", fontsize=13, fontweight="bold")
    ax.grid(True, alpha=0.3, axis="x")
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close()
    logger.info(f"Plot saved: {output_path}")


def plot_rop_depth_profile(df: pd.DataFrame, y_pred: np.ndarray, output_path: str, well_name: str = ""):
    df = df.copy()
    df["ROP_pred"] = y_pred
    df_sorted = df.sort_values("HDTH")
    
    fig, ax = plt.subplots(figsize=(6, 10))
    ax.plot(df_sorted["ROP"], df_sorted["HDTH"], color="#9FE1CB", lw=0.8, alpha=0.6, label="Actual ROP")
    ax.plot(df_sorted["ROP_pred"], df_sorted["HDTH"], color="#1D9E75", lw=1.5, alpha=0.9, label="Predicted ROP")
    ax.invert_yaxis()
    ax.set_xlabel("ROP (ft/hr)")
    ax.set_ylabel("Depth (ft)")
    ax.set_title(f"ROP vs Depth{' — ' + well_name if well_name else ''}", fontsize=12, fontweight="bold")
    ax.legend()
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close()
    logger.info(f"Plot saved: {output_path}")


def plot_optimization_results(opt_result: dict, output_path: str):
    """Bar chart of optimal vs bounds for each parameter."""
    params = opt_result["optimal_params"]
    cfg_bounds_labels = {
        "SWOB_klbs": (5, 50, "SWOB (klbs)"),
        "RPM": (40, 200, "RPM"),
        "STOR_klbf_ft": (1, 25, "Torque (klbf-ft)"),
        "CIRC_gpm": (150, 800, "Flow Rate (gpm)"),
        "SPPA_psi": (500, 3500, "Standpipe P (psi)"),
    }
    
    fig, axes = plt.subplots(1, 5, figsize=(16, 4))
    fig.suptitle(
        f"Optimized Drilling Parameters\n"
        f"Predicted ROP: {opt_result['predicted_ROP_ft_hr']:.1f} ft/hr  |  "
        f"Safety: {opt_result['safety_status']}",
        fontsize=11, fontweight="bold"
    )
    
    for ax, (key, (lo, hi, label)) in zip(axes, cfg_bounds_labels.items()):
        val = params.get(key, 0)
        pct = (val - lo) / (hi - lo) * 100
        ax.bar(["Optimal"], [pct], color="#1D9E75", width=0.4)
        ax.set_ylim(0, 100)
        ax.set_ylabel("% of range")
        ax.set_title(label, fontsize=9)
        ax.text(0, pct + 2, f"{val:.1f}", ha="center", fontsize=9, fontweight="bold")
        ax.axhline(y=100, color="#D85A30", lw=1, linestyle="--", alpha=0.5)
        ax.grid(True, alpha=0.2, axis="y")
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close()
    logger.info(f"Plot saved: {output_path}")


# ─── Excel Report ────────────────────────────────────────────

def write_excel_report(
    kpis: dict,
    robustness: dict,
    per_well_kpis: pd.DataFrame,
    fi_df: pd.DataFrame,
    opt_result: dict,
    cleaning_report: dict,
    output_path: str,
):
    wb = Workbook()
    
    _write_summary_sheet(wb.active, kpis, robustness, opt_result, cleaning_report)
    _write_kpi_sheet(wb.create_sheet("Model KPIs"), kpis, robustness)
    _write_per_well_sheet(wb.create_sheet("Per-Well Performance"), per_well_kpis)
    _write_feature_importance_sheet(wb.create_sheet("Feature Importance"), fi_df)
    _write_optimization_sheet(wb.create_sheet("Optimization Results"), opt_result)
    
    wb.save(output_path)
    logger.info(f"Excel report saved: {output_path}")


def _header_row(ws, cells: list, row: int, fill_hex: str = TEAL):
    fill = _hex(fill_hex)
    for col, val in enumerate(cells, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()


def _data_row(ws, cells: list, row: int, alt: bool = False):
    fill = _hex("E1F5EE") if alt else _hex(WHITE)
    for col, val in enumerate(cells, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
        cell.font = Font(size=10)


def _write_summary_sheet(ws, kpis, robustness, opt_result, cleaning_report):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    
    # Title
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "ROP Optimizer — Results Summary"
    c.fill = _hex(TEAL)
    c.font = Font(bold=True, color=WHITE, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    rows = [
        ("DATA CLEANING", ""),
        ("Rows before cleaning", cleaning_report.get("rows_before", "N/A")),
        ("Rows after cleaning", cleaning_report.get("rows_after", "N/A")),
        ("Rows retained (%)", f"{cleaning_report.get('pct_retained', 0):.1f}%"),
        ("Wells used", ", ".join(cleaning_report.get("wells", []))),
        ("", ""),
        ("MODEL PERFORMANCE", ""),
        ("R²", kpis.get("R2", "")),
        ("RMSE (ft/hr)", kpis.get("RMSE_ft_hr", "")),
        ("MAE (ft/hr)", kpis.get("MAE_ft_hr", "")),
        ("MAPE (%)", kpis.get("MAPE_pct", "")),
        ("Consistency Score", kpis.get("Consistency_score", "")),
        ("Hallucination Margin (%)", kpis.get("Hallucination_margin_pct", "")),
        ("Robustness Score", robustness.get("Robustness_score", "")),
        ("", ""),
        ("OPTIMIZATION RESULTS", ""),
        ("Optimal SWOB (klbs)", opt_result["optimal_params"].get("SWOB_klbs", "")),
        ("Optimal RPM", opt_result["optimal_params"].get("RPM", "")),
        ("Optimal Torque (klbf-ft)", opt_result["optimal_params"].get("STOR_klbf_ft", "")),
        ("Optimal Flow Rate (gpm)", opt_result["optimal_params"].get("CIRC_gpm", "")),
        ("Optimal SPPA (psi)", opt_result["optimal_params"].get("SPPA_psi", "")),
        ("Predicted ROP (ft/hr)", opt_result.get("predicted_ROP_ft_hr", "")),
        ("Safety Status", opt_result.get("safety_status", "")),
    ]
    
    for i, (label, val) in enumerate(rows, 2):
        if val == "" and label in ("DATA CLEANING", "MODEL PERFORMANCE", "OPTIMIZATION RESULTS"):
            ws.merge_cells(f"A{i}:B{i}")
            c = ws.cell(row=i, column=1, value=label)
            c.fill = _hex(PURPLE)
            c.font = Font(bold=True, color=WHITE, size=10)
            c.alignment = Alignment(horizontal="left", vertical="center")
        else:
            ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
            ws.cell(row=i, column=2, value=val).font = Font(size=10)


def _write_kpi_sheet(ws, kpis, robustness):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30
    
    _header_row(ws, ["KPI", "Value", "Interpretation"], 1)
    
    rows = [
        ("R²", kpis.get("R2"), "Closer to 1.0 is better"),
        ("RMSE (ft/hr)", kpis.get("RMSE_ft_hr"), "Lower is better"),
        ("MAE (ft/hr)", kpis.get("MAE_ft_hr"), "Lower is better"),
        ("MAPE (%)", kpis.get("MAPE_pct"), "Lower is better"),
        ("Consistency Score", kpis.get("Consistency_score"), "Closer to 1.0 is better"),
        ("Hallucination Margin (%)", kpis.get("Hallucination_margin_pct"), "Should be near 0%"),
        ("Directional Accuracy (%)", kpis.get("Directional_accuracy_pct"), "Higher is better"),
        ("Robustness Score", robustness.get("Robustness_score"), "Closer to 1.0 is better"),
        ("Robustness Mean Delta", robustness.get("Robustness_mean_delta_ft_hr"), "ft/hr shift under noise"),
    ]
    
    for i, (name, val, note) in enumerate(rows, 2):
        _data_row(ws, [name, val, note], i, alt=(i % 2 == 0))


def _write_per_well_sheet(ws, df: pd.DataFrame):
    ws.sheet_view.showGridLines = False
    cols = ["Well"] + list(df.columns)
    _header_row(ws, cols, 1, fill_hex=PURPLE)
    for i, (idx, row) in enumerate(df.iterrows(), 2):
        _data_row(ws, [idx] + list(row.values), i, alt=(i % 2 == 0))
    for j in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 18


def _write_feature_importance_sheet(ws, fi_df: pd.DataFrame):
    ws.sheet_view.showGridLines = False
    _header_row(ws, ["Rank", "Feature", "Importance"], 1, fill_hex=AMBER)
    for i, row in fi_df.iterrows():
        _data_row(ws, [i + 1, row["feature"], row["importance"]], i + 2, alt=(i % 2 == 0))
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 18


def _write_optimization_sheet(ws, opt_result: dict):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    
    _header_row(ws, ["Parameter", "Optimal Value"], 1, fill_hex=GREEN)
    params = opt_result["optimal_params"]
    labels = {
        "SWOB_klbs": "Weight on Bit (klbs)",
        "RPM": "Rotary Speed (RPM)",
        "STOR_klbf_ft": "Surface Torque (klbf-ft)",
        "CIRC_gpm": "Flow Rate (gpm)",
        "SPPA_psi": "Standpipe Pressure (psi)",
    }
    for i, (key, label) in enumerate(labels.items(), 2):
        _data_row(ws, [label, params.get(key, "")], i, alt=(i % 2 == 0))
    
    # Separator
    ws.cell(row=8, column=1, value="")
    _header_row(ws, ["Metric", "Value"], 9, fill_hex=TEAL)
    metrics = [
        ("Predicted ROP (ft/hr)", opt_result.get("predicted_ROP_ft_hr", "")),
        ("Safety Status", opt_result.get("safety_status", "")),
        ("Optimizer Converged", str(opt_result.get("optimizer_converged", ""))),
    ]
    if "current_ROP_ft_hr" in opt_result:
        metrics += [
            ("Current ROP (ft/hr)", opt_result.get("current_ROP_ft_hr", "")),
            ("ROP Improvement (ft/hr)", opt_result.get("ROP_improvement_ft_hr", "")),
            ("ROP Improvement (%)", opt_result.get("ROP_improvement_pct", "")),
        ]
    for i, (label, val) in enumerate(metrics, 10):
        _data_row(ws, [label, val], i, alt=(i % 2 == 0))
    
    # Safety penalties
    ws.cell(row=17, column=1, value="")
    _header_row(ws, ["Safety Check", "Penalty"], 18, fill_hex=CORAL)
    penalties = opt_result.get("safety_penalties", {})
    for i, (key, val) in enumerate(penalties.items(), 19):
        _data_row(ws, [key.replace("_", " "), round(val, 6)], i, alt=(i % 2 == 0))
